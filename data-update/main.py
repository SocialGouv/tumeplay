import openpyxl as xl
import updates


def init():
    path = 'input/input.xlsx'
    wb = xl.load_workbook(path)
    processing_xlsx(wb)


def updateContents(wb, sheet):
    wb.active = sheet
    ws = wb.active
    for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
        if row[0] is None:
            break
        data = {
            "title": row[0],
            "new_title": row[1],
            "content": row[2],
            "theme": row[3],
            "theme_app": row[4],
            "etiquette": row[5],
            "level": row[6]
        }
        updates.updateContent(data)


def updateQuestions(wb, sheet):
    wb.active = sheet
    ws = wb.active
    for row in ws.iter_rows(min_row=2, max_col=16, values_only=True):
        if row[0] is None:
            break

        data = {
            "title": row[0],
            "title_related_content": row[9],
            "kind": row[11]
        }

        if (row[11] == "Trou"):
            data["newTitle"] = row[12]
            splitted_responses = row[13].split("\n")
            data["response_A_mobile"] = splitted_responses[0].replace('- ', '')
            data["response_B_mobile"] = splitted_responses[1].replace('- ', '')
            data["response_C_mobile"] = splitted_responses[2].replace('- ', '')

        updates.updateQuestion(data)

def updateModules(wb, sheet):
    wb.active = sheet
    ws = wb.active


def processing_xlsx(wb):
    for i, s in enumerate(wb.sheetnames):
        if s == 'contents':
            contents_sheet = i
        if s == 'questions':
            questions_sheet = i
        if s == 'MODULES':
            modules_sheeet = i

    updateContents(wb, contents_sheet)
    updateQuestions(wb, questions_sheet)
    # updateModules(wb, modules_sheeet)

init()